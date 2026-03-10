"""
ECU-tab — overzicht ECU 1 / ECU 2 per vliegtuig.
Selectie via dropdown, tabel conform Planning-tab opmaak.
"""
import pandas as pd
from PySide6.QtCore import Qt, QSize, QTimer
from PySide6.QtGui import QIcon
from PySide6.QtWidgets import (
    QAbstractItemView, QComboBox, QFrame, QHBoxLayout, QHeaderView,
    QLabel, QPushButton, QSizePolicy, QTableView, QVBoxLayout, QWidget,
)

from ui.theme import SLATE_600, SLATE_700, SLATE_800, SLATE_900, WHITE
from ui.tabs.planning_tab import (
    PlanningTableModel, MultiColumnFilterProxy, FilterHeaderView,
    _FilterPopup, _make_clear_filter_icon, _EXPORT_BTN_QSS, PLAN_ROW_H,
    _CompactItemDelegate,
)
from ui.tabs.overview_tab import _TBL_QSS, TBL_BG, TBL_BDR, TBL_TEXT

# ── Style ─────────────────────────────────────────────────────────────────
_CTRL_QSS = f"""
    QLabel   {{ color: {WHITE}; font-size: 11px; background: transparent; }}
    QComboBox {{
        background-color: #d0d0d0; color: #1a1a1a;
        border: 1px solid {SLATE_600}; border-radius: 3px;
        padding: 1px 4px; font-size: 12px; max-height: 20px;
    }}
    QComboBox::drop-down {{ border: none; width: 18px; }}
    QComboBox QAbstractItemView {{
        font-size: 13px; background-color: #d0d0d0; color: #1a1a1a;
        selection-background-color: #a0a8b0;
    }}
"""
_TOGGLE_ACTIVE = (
    'QPushButton { background: #1d4ed8; color: white; '
    'border: 1px solid #3b82f6; border-radius: 0; '
    'padding: 3px 20px; font-size: 11px; font-weight: bold; }'
)
_TOGGLE_INACTIVE = (
    'QPushButton { background: #334155; color: #94a3b8; '
    'border: 1px solid #475569; border-radius: 0; '
    'padding: 3px 20px; font-size: 11px; }'
    'QPushButton:hover { background: #3e5068; color: white; }'
)


class EcuTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._ac_list:    list = []
        self._df_cyc            = None
        self._df_cfg            = None
        self._sys_vars:   dict = {}
        self._active_ecu: str  = '1'
        self._last_col_sig: tuple | None = None
        self._refresh_timer = QTimer(self)
        self._refresh_timer.setSingleShot(True)
        self._refresh_timer.timeout.connect(self._do_refresh)
        self._build_ui()

    # ── UI ────────────────────────────────────────────────────────────────
    def _build_ui(self) -> None:
        self.setStyleSheet(f'background: {SLATE_900};')
        main = QVBoxLayout(self)
        main.setContentsMargins(8, 6, 6, 6)
        main.setSpacing(4)

        # ── Controlebar ───────────────────────────────────────────────────
        ctrl_frame = QFrame()
        ctrl_frame.setStyleSheet(
            f'QFrame {{ background: {SLATE_800}; border-radius: 4px; border: 1px solid {SLATE_600}; }}'
            + _CTRL_QSS
        )
        ctrl_row = QHBoxLayout(ctrl_frame)
        ctrl_row.setContentsMargins(10, 6, 10, 6)
        ctrl_row.setSpacing(12)

        # Aircraft label + combo
        ac_col = QVBoxLayout()
        ac_col.setSpacing(2)
        ac_lbl = QLabel('Aircraft')
        ac_lbl.setStyleSheet(f'font-size: 12px; color: {WHITE};')
        ac_col.addWidget(ac_lbl)
        self._combo_ac = QComboBox()
        self._combo_ac.setFixedWidth(95)
        self._combo_ac.setMaxVisibleItems(30)
        self._combo_ac.setItemDelegate(_CompactItemDelegate(self._combo_ac))
        self._combo_ac.currentTextChanged.connect(self._schedule_refresh)
        ac_col.addWidget(self._combo_ac)
        ctrl_row.addLayout(ac_col)

        # ECU toggle
        ecu_col = QVBoxLayout()
        ecu_col.setSpacing(2)
        ecu_lbl = QLabel('ECU')
        ecu_lbl.setStyleSheet(f'font-size: 12px; color: {WHITE};')
        ecu_col.addWidget(ecu_lbl)
        toggle_row = QHBoxLayout()
        toggle_row.setSpacing(0)
        self._btn_ecu1 = QPushButton('ECU 1')
        self._btn_ecu2 = QPushButton('ECU 2')
        self._btn_ecu1.setFixedHeight(20)
        self._btn_ecu2.setFixedHeight(20)
        self._btn_ecu1.clicked.connect(lambda: self._switch_ecu('1'))
        self._btn_ecu2.clicked.connect(lambda: self._switch_ecu('2'))
        toggle_row.addWidget(self._btn_ecu1)
        toggle_row.addWidget(self._btn_ecu2)
        ecu_col.addLayout(toggle_row)
        ctrl_row.addLayout(ecu_col)

        ctrl_row.addStretch()
        main.addWidget(ctrl_frame)
        main.addSpacing(6)

        # ── Info-balk ─────────────────────────────────────────────────────
        info_bar = QWidget()
        info_bar.setFixedHeight(38)
        info_bar.setStyleSheet(f'background: {SLATE_800}; border-radius: 4px;')
        ib = QHBoxLayout(info_bar)
        ib.setContentsMargins(10, 0, 10, 0)
        ib.setSpacing(10)

        self._lbl_info = QLabel('Selecteer een vliegtuig')
        self._lbl_info.setStyleSheet(f'color: {WHITE}; font-size: 11px; background: transparent;')
        ib.addWidget(self._lbl_info)
        ib.addSpacing(40)

        self._btn_clear = QPushButton()
        self._btn_clear.setIcon(_make_clear_filter_icon(16))
        self._btn_clear.setToolTip('Verwijder alle filters')
        self._btn_clear.setFixedSize(22, 22)
        self._btn_clear.setStyleSheet(
            'QPushButton { background: transparent; border: none; }'
            'QPushButton:hover { background: rgba(255,255,255,30); border-radius: 3px; }'
        )
        ib.addWidget(self._btn_clear)
        ib.addStretch()

        self._btn_export = QPushButton('')
        self._btn_export.setStyleSheet(_EXPORT_BTN_QSS)
        self._btn_export.setFixedSize(40, 30)
        self._btn_export.setEnabled(False)
        self._btn_export.clicked.connect(self._export_excel)
        try:
            import tempfile, os
            from PySide6.QtWidgets import QFileIconProvider
            from PySide6.QtCore import QFileInfo
            _t = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
            _t.close()
            self._btn_export.setIcon(QFileIconProvider().icon(QFileInfo(_t.name)))
            self._btn_export.setIconSize(QSize(24, 24))
            os.unlink(_t.name)
        except Exception:
            pass
        ib.addWidget(self._btn_export)

        # ── Tabel ─────────────────────────────────────────────────────────
        self._model = PlanningTableModel(self)
        self._proxy = MultiColumnFilterProxy(self)
        self._proxy.setSourceModel(self._model)
        self._proxy.filtersChanged.connect(self._update_info)
        self._btn_clear.clicked.connect(self._proxy.clear_all_filters)

        self._tbl = QTableView()
        self._tbl.setStyleSheet(
            _TBL_QSS
            + f'QTableView {{ color: {TBL_TEXT}; gridline-color: {TBL_BDR}; font-size: 11px; }}'
            + f'QTableView::item {{ color: {TBL_TEXT}; background-color: {TBL_BG}; '
              f'font-size: 11px; padding-left: 4px; padding-right: 4px; }}'
        )
        self._tbl.setModel(self._proxy)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.verticalHeader().setDefaultSectionSize(PLAN_ROW_H)
        self._tbl.verticalHeader().setMinimumSectionSize(1)
        self._tbl.verticalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Fixed)
        self._tbl.setWordWrap(False)
        self._tbl.setAlternatingRowColors(False)
        self._tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tbl.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self._tbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        self._tbl.setSortingEnabled(True)

        self._hdr = FilterHeaderView(self._proxy, self._tbl)
        self._hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)
        self._hdr.setFixedHeight(28)
        self._tbl.setHorizontalHeader(self._hdr)
        self._hdr.filterArrowClicked.connect(self._show_filter_popup)

        right_col = QVBoxLayout()
        right_col.setContentsMargins(50, 0, 0, 0)
        right_col.setSpacing(4)
        right_col.addWidget(info_bar)
        right_col.addWidget(self._tbl, stretch=1)
        main.addLayout(right_col, stretch=1)

        self._set_toggle(self._active_ecu)

    # ── Data laden ────────────────────────────────────────────────────────
    def load_data(self, ac_list: list, df_cyc, df_cfg, sys_vars: dict) -> None:
        self._ac_list  = ac_list or []
        self._df_cyc   = df_cyc
        self._df_cfg   = df_cfg
        self._sys_vars = sys_vars
        current = self._combo_ac.currentText()
        self._combo_ac.blockSignals(True)
        self._combo_ac.clear()
        self._combo_ac.addItems(self._ac_list)
        if current in self._ac_list:
            self._combo_ac.setCurrentText(current)
        self._combo_ac.blockSignals(False)
        self._schedule_refresh()

    # ── ECU toggle ────────────────────────────────────────────────────────
    def _switch_ecu(self, nr: str) -> None:
        self._active_ecu = nr
        self._set_toggle(nr)
        self._schedule_refresh()

    def _set_toggle(self, nr: str) -> None:
        self._btn_ecu1.setStyleSheet(_TOGGLE_ACTIVE if nr == '1' else _TOGGLE_INACTIVE)
        self._btn_ecu2.setStyleSheet(_TOGGLE_ACTIVE if nr == '2' else _TOGGLE_INACTIVE)

    # ── Refresh ───────────────────────────────────────────────────────────
    def _schedule_refresh(self, _=None) -> None:
        self._refresh_timer.start(20)

    def _do_refresh(self) -> None:
        aircraft = self._combo_ac.currentText()
        if not aircraft or self._df_cyc is None:
            return
        from data.processor import get_ecu_status, get_ecu_serienumber
        try:
            df = get_ecu_status(aircraft, self._active_ecu, self._df_cyc, self._sys_vars)
            sn = (
                get_ecu_serienumber(aircraft, self._active_ecu, self._df_cfg)
                if self._df_cfg is not None else 'n/b'
            )
        except Exception as exc:
            self._lbl_info.setText(f'Fout: {exc}')
            return
        self._populate_table(df, aircraft, sn)

    def _populate_table(self, df: pd.DataFrame, aircraft: str, sn: str) -> None:
        self._proxy.clear_all_filters()
        self._model.load(df)
        self._apply_row_heights()
        QTimer.singleShot(0, self._apply_row_heights)

        if df.empty:
            self._lbl_info.setText(f'{aircraft}  —  ECU {self._active_ecu}  —  SN {sn}  •  Geen data')
            self._btn_export.setEnabled(False)
            return

        sig = tuple(df.columns)
        if sig != self._last_col_sig:
            QTimer.singleShot(0, self._tbl.resizeColumnsToContents)
            self._last_col_sig = sig

        self._btn_export.setEnabled(True)
        self._update_info(aircraft=aircraft, sn=sn)

    def _apply_row_heights(self) -> None:
        self._tbl.verticalHeader().setDefaultSectionSize(PLAN_ROW_H)
        for r in range(self._proxy.rowCount()):
            self._tbl.setRowHeight(r, PLAN_ROW_H)

    def _update_info(self, aircraft: str = '', sn: str = '') -> None:
        if not aircraft:
            aircraft = self._combo_ac.currentText()
        visible = self._proxy.rowCount()
        total   = self._model.rowCount()
        base    = f'{aircraft}  —  ECU {self._active_ecu}  —  SN {sn}' if sn else aircraft
        if visible == total:
            self._lbl_info.setText(f'{base}  •  {total} regel{"s" if total != 1 else ""}')
        else:
            self._lbl_info.setText(f'{base}  •  {visible} van {total} regels zichtbaar')

    # ── Filter popup ──────────────────────────────────────────────────────
    def _show_filter_popup(self, col: int, pos) -> None:
        popup = _FilterPopup(
            self._proxy.unique_values(col),
            self._proxy.current_filter(col),
            self,
        )
        popup.applied.connect(lambda sel, c=col: self._proxy.set_filter(c, sel))
        popup.sorted.connect(lambda order, c=col: self._tbl.sortByColumn(c, Qt.SortOrder(order)))
        popup.adjustSize()
        popup.move(pos.x() - popup.width() // 2, pos.y() + 4)
        popup.show()

    # ── Export ────────────────────────────────────────────────────────────
    def _export_excel(self) -> None:
        aircraft = self._combo_ac.currentText()
        if not aircraft or self._df_cyc is None:
            return
        import os, tempfile
        from data.processor import get_ecu_status, get_ecu_serienumber
        try:
            df1 = get_ecu_status(aircraft, '1', self._df_cyc, self._sys_vars)
            df2 = get_ecu_status(aircraft, '2', self._df_cyc, self._sys_vars)
            sn1 = get_ecu_serienumber(aircraft, '1', self._df_cfg) if self._df_cfg is not None else 'n/b'
            sn2 = get_ecu_serienumber(aircraft, '2', self._df_cfg) if self._df_cfg is not None else 'n/b'
        except Exception as exc:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, 'Export fout', str(exc))
            return
        tab1 = f'{aircraft} ECU1 SN {sn1}'[:31]
        tab2 = f'{aircraft} ECU2 SN {sn2}'[:31]
        try:
            tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', prefix=f'ECU_{aircraft}_', delete=False)
            tmp.close()
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                df1.to_excel(writer, index=False, sheet_name=tab1)
                df2.to_excel(writer, index=False, sheet_name=tab2)
                _style_excel(writer.sheets[tab1], df1)
                _style_excel(writer.sheets[tab2], df2)
            os.startfile(tmp.name)
        except Exception as exc:
            from PySide6.QtWidgets import QMessageBox
            QMessageBox.warning(self, 'Export fout', str(exc))


def _style_excel(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    _HDR_FILL = PatternFill('solid', fgColor='AFC4D0')
    _ALT_FILL = PatternFill('solid', fgColor='EEF2FA')
    _WHT_FILL = PatternFill('solid', fgColor='FFFFFF')
    _HDR_FONT = Font(name='Calibri', bold=True, size=10, color='0D1A26')
    _DAT_FONT = Font(name='Calibri', size=10, color='1E293B')
    _THIN = Side(style='thin', color='C8D4E8')
    _BDR  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
    for i, col in enumerate(df.columns, start=1):
        lengths = [len(str(col))] + [len(str(v)) for v in df[col] if pd.notna(v)]
        ws.column_dimensions[get_column_letter(i)].width = min(max(max(lengths) + 2, 10), 50)
    for cell in ws[1]:
        cell.fill = _HDR_FILL; cell.font = _HDR_FONT; cell.border = _BDR
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        ws.row_dimensions[r_idx + 2].height = 14
        fill = _WHT_FILL if r_idx % 2 == 0 else _ALT_FILL
        for cell in row:
            cell.font = _DAT_FONT; cell.border = _BDR; cell.fill = fill
            cell.alignment = Alignment(vertical='center')
    ws.freeze_panes = 'A2'
