"""
Overzichtscherm voor MV3 — 4 tabellen naast elkaar per vliegtuig.
Linkerkolom: kist-nr, vlieguren, detail-knoppen.
"""
import pandas as pd
from PySide6.QtCore import Qt, QEvent, QObject, QSize
from PySide6.QtGui import QBrush, QColor, QFont, QIcon
from PySide6.QtWidgets import (
    QAbstractItemView, QDialog, QFrame, QHBoxLayout, QHeaderView, QLabel, QMessageBox,
    QPushButton, QScrollArea, QSizePolicy, QStyledItemDelegate,
    QTableWidget, QTableWidgetItem, QVBoxLayout, QWidget,
)


class _BgDelegate(QStyledItemDelegate):
    """Zorgt dat setBackground() zichtbaar blijft ook met een stylesheet."""
    def paint(self, painter, option, index):
        bg = index.data(Qt.ItemDataRole.BackgroundRole)
        if bg is not None:
            brush = bg if isinstance(bg, QBrush) else QBrush(bg)
            painter.fillRect(option.rect, brush)
        # Teken alleen tekst/icoon, niet de style-achtergrond opnieuw
        opt = type(option)(option)
        opt.backgroundBrush = QBrush()  # leeg = transparant voor super
        super().paint(painter, opt, index)

from ui.theme import SLATE_400, SLATE_600, SLATE_700, SLATE_800, SLATE_900, WHITE

# ---------------------------------------------------------------------------
# Constanten
# ---------------------------------------------------------------------------
N_ROWS       = 14
ROW_H        = 10
HDR_H        = 28       # 2-regelige kolomkop
TBL_H        = N_ROWS * ROW_H + HDR_H + 4   # ~172 px

AC_CARD_COLOR = '#1c3e72'   # staal blauw  — info-kaart & blok-rand
AC_CARD_BDR   = '#2b54a0'   # staal blauw rand
AC_COL_HDR    = '#afc4d0'   # licht grijs-blauw — kolomkoppen
AC_COL_BDR    = '#8aafc0'   # rand kolomkoppen

# backwards-compat alias (gebruikt in _build_info_card via de kaart-border)
AC_HDR_COLOR  = AC_CARD_COLOR
AC_HDR_BDR    = AC_CARD_BDR
TBL_BG       = '#ffffff'
TBL_ALT      = '#eef2fa'
TBL_TEXT     = '#1e293b'
TBL_BDR      = '#c8d4e8'
NEG_BG       = '#f4a0a0'   # zacht rood
NEG_FG       = '#1e293b'   # zwart

INFO_W       = 88       # breedte info-kaart
POPLAN_W     = 48       # breedte Po Plan-kolom
POREF_W      = 72       # breedte Po Ref-kolom
OMSCHR_MAX   = 180      # max breedte Omschrijving-kolom (dynamisch/stretch)

# ---------------------------------------------------------------------------
# QSS
# ---------------------------------------------------------------------------
_TBL_QSS = f"""
    QTableWidget {{
        gridline-color: {TBL_BDR};
        color: {TBL_TEXT};
        font-size: 10px;
        border: none;
        selection-background-color: #b8ccee;
        selection-color: {TBL_TEXT};
    }}
    QHeaderView::section {{
        background-color: {AC_COL_HDR};
        color: #0d1a26;
        font-weight: bold;
        font-size: 10px;
        border: none;
        border-right: 1px solid {AC_COL_BDR};
        padding: 2px 4px;
    }}
    QTableWidget::item {{
        padding: 0px 2px;
        border-bottom: 1px solid {TBL_BDR};
    }}
    QScrollBar:vertical {{
        background: #e8edf5; width: 7px; border-radius: 3px;
    }}
    QScrollBar::handle:vertical {{
        background: #94a3b8; border-radius: 3px; min-height: 16px;
    }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
"""

_BTN_QSS = f"""
    QPushButton {{
        background-color: qlineargradient(
            x1:0, y1:0, x2:0, y2:1,
            stop:0 {AC_COL_BDR}, stop:1 {AC_COL_HDR}
        );
        color: #1a1a1a;
        border: 1px solid {AC_COL_BDR};
        border-bottom: 2px solid #3a5a6e;
        border-radius: 4px;
        font-size: 10px;
        font-weight: normal;
        padding: 1px 6px;
    }}
    QPushButton:hover {{
        background-color: qlineargradient(
            x1:0, y1:0, x2:0, y2:1,
            stop:0 #8aabb8, stop:1 {AC_COL_BDR}
        );
    }}
    QPushButton:pressed {{
        background-color: {AC_COL_HDR};
        border-bottom: 1px solid #3a5a6e;
        padding-top: 5px;
    }}
"""

_DLG_QSS = f"""
    QDialog {{ background-color: {SLATE_900}; color: {WHITE}; font-size: 13px; }}
    QLabel  {{ color: {WHITE}; background: transparent; }}
    QTableWidget {{
        background-color: {TBL_BG};
        gridline-color: {TBL_BDR};
        color: {TBL_TEXT};
        font-size: 12px;
        border: 1px solid {TBL_BDR};
        selection-background-color: #b8ccee;
        selection-color: {TBL_TEXT};
    }}
    QHeaderView::section {{
        background-color: {AC_COL_HDR};
        color: #1a2530;
        font-weight: bold;
        font-size: 11px;
        border: none;
        border-right: 1px solid {AC_COL_BDR};
        padding: 4px 8px;
    }}
    QTableWidget::item {{
        padding: 2px 8px;
        border-bottom: 1px solid {TBL_BDR};
    }}
    QScrollBar:vertical {{
        background: #e8edf5; width: 8px; border-radius: 4px;
    }}
    QScrollBar::handle:vertical {{
        background: #94a3b8; border-radius: 4px; min-height: 20px;
    }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
"""

_DLG_ROW_H = 26
_DLG_HDR_H = 36


def _style_excel_sheet(ws, df: pd.DataFrame) -> None:
    """Pas uniforme opmaak toe op een openpyxl worksheet."""
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    _HDR_FILL = PatternFill('solid', fgColor='AFC4D0')
    _ALT_FILL = PatternFill('solid', fgColor='EEF2FA')
    _WHT_FILL = PatternFill('solid', fgColor='FFFFFF')
    _HDR_FONT = Font(name='Calibri', bold=True, size=10, color='0D1A26')
    _DAT_FONT = Font(name='Calibri', size=10, color='1E293B')
    _THIN = Side(style='thin', color='C8D4E8')
    _BDR  = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

    # Kolombreedte op basis van inhoud
    for i, col in enumerate(df.columns, start=1):
        lengths = [len(str(col))] + [len(str(v)) for v in df[col] if pd.notna(v)]
        ws.column_dimensions[get_column_letter(i)].width = min(max(max(lengths) + 2, 10), 50)

    # Header
    for cell in ws[1]:
        cell.fill = _HDR_FILL
        cell.font = _HDR_FONT
        cell.border = _BDR
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 22

    # Data rijen
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        ws.row_dimensions[r_idx + 2].height = 14
        fill = _WHT_FILL if r_idx % 2 == 0 else _ALT_FILL
        for cell in row:
            cell.font = _DAT_FONT
            cell.border = _BDR
            cell.fill = fill
            cell.alignment = Alignment(vertical='center')

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions


def _open_df_in_excel(df: pd.DataFrame, title: str, parent) -> None:
    import os, tempfile
    try:
        safe = ''.join(c if c.isalnum() or c in '-_ ' else '_' for c in title)
        sheet = safe[:31]
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', prefix=f'{safe[:20]}_', delete=False)
        tmp.close()
        with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name=sheet)
            _style_excel_sheet(writer.sheets[sheet], df)
        os.startfile(tmp.name)
    except Exception as exc:
        QMessageBox.warning(parent, 'Export error', str(exc))


def _excel_icon() -> QIcon:
    import os, tempfile
    from PySide6.QtWidgets import QFileIconProvider
    from PySide6.QtCore import QFileInfo
    try:
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp.close()
        icon = QFileIconProvider().icon(QFileInfo(tmp.name))
        os.unlink(tmp.name)
        return icon
    except Exception:
        return QIcon()


# ---------------------------------------------------------------------------
# Dialogen
# ---------------------------------------------------------------------------

class _InfoDialog(QDialog):
    def __init__(self, title: str, df: pd.DataFrame, col_widths: list, parent=None, width: int = 460):
        super().__init__(parent)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint
        )
        self.setWindowTitle(title)
        self.setStyleSheet(_DLG_QSS)
        self.setMinimumSize(420, 380)
        self.resize(width, 440)
        self._df    = df
        self._title = title

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        lbl = QLabel(f'<b style="font-size:14px">{title}</b>')
        layout.addWidget(lbl)

        tbl = _build_table(df, list(df.columns), list(df.columns),
                           col_widths=col_widths, rest_col=None, max_rows=60)
        # Overschrijf fixed height zodat de tabel meegroeit met het venster
        tbl.setMinimumHeight(300)
        tbl.setMaximumHeight(16777215)
        tbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        tbl.horizontalHeader().setFixedHeight(_DLG_HDR_H)
        for r in range(tbl.rowCount()):
            tbl.setRowHeight(r, _DLG_ROW_H)
        layout.addWidget(tbl, stretch=1)

        btn_row = QHBoxLayout()
        btn_xls = QPushButton(_excel_icon(), '')
        btn_xls.setStyleSheet(_BTN_QSS)
        btn_xls.setFixedSize(36, 32)
        btn_xls.setIconSize(QSize(24, 24))
        btn_xls.clicked.connect(lambda: _open_df_in_excel(self._df, self._title, self))
        btn_row.addWidget(btn_xls)
        btn_row.addStretch()
        btn_close = QPushButton('Close')
        btn_close.setStyleSheet(_BTN_QSS)
        btn_close.setFixedHeight(32)
        btn_close.setFixedWidth(90)
        btn_close.clicked.connect(self.accept)
        btn_row.addWidget(btn_close)
        layout.addLayout(btn_row)


# ---------------------------------------------------------------------------
# Specials-dialoog (bewerkbaar)
# ---------------------------------------------------------------------------

class _SpecialsDialog(QDialog):
    _DATA_COLS = ['Bijzonderheid', 'DueDate', 'Eenheid', 'PoPlan', 'UitvoerenBij']
    _HEADERS   = ['Bijzonderheid', 'Due Date', 'Eenheid', 'PO Plan', 'Uitvoeren bij', '']
    _DEL_COL   = 5   # index van de prullenbak-kolom

    _DEL_BTN_QSS = (
        'QPushButton { color: #c0392b; font-weight: bold; font-size: 14px; '
        'border: 1px solid #c0392b; border-radius: 3px; background: transparent; padding: 0; }'
        'QPushButton:hover { background: #c0392b; color: white; }'
    )

    def __init__(self, aircraft: str, user_vars: dict, parent=None):
        super().__init__(parent)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint
        )
        self.setWindowTitle(f'Specials — {aircraft}')
        self.setStyleSheet(_DLG_QSS)
        self.setMinimumSize(700, 380)
        self.resize(780, 460)

        self._aircraft  = aircraft
        self._user_vars = user_vars
        self._is_dirty  = False

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        lbl = QLabel(f'<b style="font-size:14px">Specials — {aircraft}</b>')
        layout.addWidget(lbl)

        self._tbl = QTableWidget(0, len(self._HEADERS))
        self._tbl.setHorizontalHeaderLabels(self._HEADERS)
        self._tbl.verticalHeader().setVisible(False)
        self._tbl.horizontalHeader().setFixedHeight(32)
        self._tbl.setMinimumHeight(260)
        self._tbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        hdr = self._tbl.horizontalHeader()
        hdr.setSectionResizeMode(0, QHeaderView.ResizeMode.Stretch)
        for i in (1, 2, 3, 4):
            hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
        hdr.setSectionResizeMode(self._DEL_COL, QHeaderView.ResizeMode.Fixed)
        self._tbl.setColumnWidth(1, 90)
        self._tbl.setColumnWidth(2, 80)
        self._tbl.setColumnWidth(3, 70)
        self._tbl.setColumnWidth(4, 90)
        self._tbl.setColumnWidth(self._DEL_COL, 30)

        layout.addWidget(self._tbl, stretch=1)

        btns = QHBoxLayout()
        btn_xls = QPushButton(_excel_icon(), '')
        btn_rst = QPushButton('Reset')
        btn_sav = QPushButton('Save && close')
        self._reset_button = btn_rst
        for b in (btn_rst, btn_sav):
            b.setStyleSheet(_BTN_QSS)
            b.setFixedHeight(28)
        btn_xls.setStyleSheet(_BTN_QSS)
        btn_xls.setFixedSize(32, 28)
        btn_xls.setIconSize(QSize(20, 20))
        btn_xls.clicked.connect(self._export_excel)
        btn_rst.clicked.connect(lambda checked: self._on_reset())
        btn_sav.clicked.connect(self._save)
        btns.addWidget(btn_xls)
        btns.addSpacing(8)
        btns.addWidget(btn_rst)
        btns.addStretch()
        btns.addWidget(btn_sav)
        layout.addLayout(btns)

        self._load_data()
        self._tbl.cellChanged.connect(self._on_cell_changed)

    def _load_data(self):
        bijz = (self._user_vars
                .get('helikopter', {})
                .get(self._aircraft, {})
                .get('InspBijzonderheden', {}))
        self._tbl.blockSignals(True)
        self._tbl.setRowCount(0)
        for _, item in bijz.items():
            if any(item.get(col, '') for col in self._DATA_COLS):
                self._append_data_row(item)
        self._append_empty_row()
        self._tbl.blockSignals(False)
        self._is_dirty = False
        self._reset_button.setEnabled(False)

    def _append_data_row(self, item: dict):
        row = self._tbl.rowCount()
        self._tbl.insertRow(row)
        for c, col in enumerate(self._DATA_COLS):
            self._tbl.setItem(row, c, QTableWidgetItem(item.get(col, '')))
        self._tbl.setRowHeight(row, _DLG_ROW_H)
        self._set_delete_btn(row)

    def _append_empty_row(self):
        row = self._tbl.rowCount()
        self._tbl.insertRow(row)
        for c in range(len(self._DATA_COLS)):
            self._tbl.setItem(row, c, QTableWidgetItem(''))
        self._tbl.setRowHeight(row, _DLG_ROW_H)

    def _set_delete_btn(self, row: int):
        btn = QPushButton('x')
        btn.setFixedSize(24, 20)
        btn.setStyleSheet(self._DEL_BTN_QSS)
        btn.clicked.connect(self._on_delete_row)
        self._tbl.setCellWidget(row, self._DEL_COL, btn)

    def _on_delete_row(self):
        btn = self.sender()
        for r in range(self._tbl.rowCount()):
            if self._tbl.cellWidget(r, self._DEL_COL) is btn:
                self._tbl.removeRow(r)
                self._set_dirty()
                break

    def _on_cell_changed(self, row: int, col: int):
        self._set_dirty()
        if row == self._tbl.rowCount() - 1:
            # lege invoerregel is ingevuld -> prullenbak toevoegen + nieuwe lege regel
            self._set_delete_btn(row)
            self._tbl.blockSignals(True)
            self._append_empty_row()
            self._tbl.blockSignals(False)

    def _save(self):
        from data.processor import save_user_variables, load_user_variables
        bijz = {}
        idx = 0
        for r in range(self._tbl.rowCount()):
            item = {}
            for c, col in enumerate(self._DATA_COLS):
                cell = self._tbl.item(r, c)
                item[col] = cell.text().strip() if cell else ''
            if any(item.values()):
                bijz[str(idx)] = item
                idx += 1

        # Herlaad de laatste versie van schijf zodat andere vliegtuigen niet overschreven worden
        try:
            current = load_user_variables()
        except Exception:
            current = self._user_vars

        # Detecteer conflict: zijn de specials van dit vliegtuig gewijzigd door iemand anders?
        original_bijz = (self._user_vars
                         .get('helikopter', {})
                         .get(self._aircraft, {})
                         .get('InspBijzonderheden', {}))
        current_bijz  = (current
                         .get('helikopter', {})
                         .get(self._aircraft, {})
                         .get('InspBijzonderheden', {}))
        if current_bijz != original_bijz:
            reply = QMessageBox.warning(
                self, 'Conflict detected',
                f'Specials for {self._aircraft} were modified by another user since you opened this dialog.\n\n'
                'Do you want to overwrite their changes with yours?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            )
            if reply != QMessageBox.StandardButton.Yes:
                return

        # Schrijf alleen de specials van dit vliegtuig in de vers herladen data
        (current
         .setdefault('helikopter', {})
         .setdefault(self._aircraft, {})
         )['InspBijzonderheden'] = bijz
        save_user_variables(current)
        self.accept()

    def _export_excel(self):
        rows = []
        for r in range(self._tbl.rowCount()):
            row = {col: (self._tbl.item(r, c).text() if self._tbl.item(r, c) else '')
                   for c, col in enumerate(self._DATA_COLS)}
            if any(row.values()):
                rows.append(row)
        _open_df_in_excel(pd.DataFrame(rows), f'Specials_{self._aircraft}', self)

    def _on_reset(self):
        if not self._is_dirty:
            return
        if QMessageBox.question(self, 'Reset specials',
                                'Discard unsaved changes?',
                                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No) == QMessageBox.StandardButton.Yes:
            self._load_data()

    def _set_dirty(self):
        self._is_dirty = True
        self._reset_button.setEnabled(True)

# ---------------------------------------------------------------------------
# ECU-dialoog (ECU 1 / ECU 2 switcher)
# ---------------------------------------------------------------------------

_ECU_COL_WIDTHS = [('F', 80), ('F', 110), ('S', 0), ('F', 90), ('F', 70)]

_TOGGLE_ACTIVE = (
    'QPushButton { background: #1d4ed8; color: white; border: 1px solid #3b82f6; '
    'padding: 3px 16px; font-size: 11px; font-weight: bold; border-radius: 0; }'
)
_TOGGLE_INACTIVE = (
    'QPushButton { background: #334155; color: #94a3b8; border: 1px solid #475569; '
    'padding: 3px 16px; font-size: 11px; border-radius: 0; }'
    'QPushButton:hover { background: #3e5068; color: white; }'
)


class _EcuDialog(QDialog):
    def __init__(self, aircraft: str, df_cyc, df_cfg, sys_vars: dict, parent=None):
        super().__init__(parent)
        self.setWindowFlags(
            Qt.WindowType.Dialog |
            Qt.WindowType.WindowTitleHint |
            Qt.WindowType.WindowCloseButtonHint
        )
        self.setWindowTitle(f'ECU  —  {aircraft}')
        self.setStyleSheet(_DLG_QSS)
        self.setMinimumSize(500, 380)
        self.resize(640, 460)

        self._aircraft   = aircraft
        self._df_cyc     = df_cyc
        self._df_cfg     = df_cfg
        self._sys_vars   = sys_vars
        self._current_df = pd.DataFrame()
        self._current_ecu = '1'

        layout = QVBoxLayout(self)
        layout.setContentsMargins(16, 16, 16, 16)
        layout.setSpacing(10)

        # Toggle-rij
        toggle_row = QHBoxLayout()
        toggle_row.setSpacing(0)
        self._btn1 = QPushButton('ECU 1')
        self._btn2 = QPushButton('ECU 2')
        self._btn1.clicked.connect(lambda: self._switch('1'))
        self._btn2.clicked.connect(lambda: self._switch('2'))
        toggle_row.addWidget(self._btn1)
        toggle_row.addWidget(self._btn2)
        toggle_row.addStretch()
        layout.addLayout(toggle_row)

        self._title_lbl = QLabel()
        layout.addWidget(self._title_lbl)

        self._tbl_holder = QWidget()
        self._tbl_layout = QVBoxLayout(self._tbl_holder)
        self._tbl_layout.setContentsMargins(0, 0, 0, 0)
        layout.addWidget(self._tbl_holder, stretch=1)

        btn_row = QHBoxLayout()
        btn_xls = QPushButton(_excel_icon(), '')
        btn_xls.setStyleSheet(_BTN_QSS)
        btn_xls.setFixedSize(36, 32)
        btn_xls.setIconSize(QSize(24, 24))
        btn_xls.clicked.connect(self._export_both)
        btn_row.addWidget(btn_xls)
        btn_row.addStretch()
        close_btn = QPushButton('Close')
        close_btn.setStyleSheet(_BTN_QSS)
        close_btn.setFixedHeight(32)
        close_btn.setFixedWidth(90)
        close_btn.clicked.connect(self.accept)
        btn_row.addWidget(close_btn)
        layout.addLayout(btn_row)

        self._switch('1')

    def _switch(self, ecu_nr: str):
        import traceback
        from data.processor import get_ecu_status, get_ecu_serienumber

        self._btn1.setStyleSheet(_TOGGLE_ACTIVE if ecu_nr == '1' else _TOGGLE_INACTIVE)
        self._btn2.setStyleSheet(_TOGGLE_ACTIVE if ecu_nr == '2' else _TOGGLE_INACTIVE)

        try:
            df = get_ecu_status(self._aircraft, ecu_nr, self._df_cyc, self._sys_vars)
            sn = get_ecu_serienumber(self._aircraft, ecu_nr, self._df_cfg) if self._df_cfg is not None else 'n/b'
        except Exception:
            traceback.print_exc()
            return

        self._current_df  = df
        self._current_ecu = ecu_nr

        self._title_lbl.setText(f'<b style="font-size:13px">{self._aircraft}  ECU{ecu_nr}  SN {sn}</b>')

        # Verwijder oude tabel
        while self._tbl_layout.count():
            item = self._tbl_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        tbl = _build_table(df, list(df.columns), list(df.columns),
                           col_widths=_ECU_COL_WIDTHS, rest_col=None, max_rows=60)
        tbl.setMinimumHeight(300)
        tbl.setMaximumHeight(16777215)
        tbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)
        tbl.horizontalHeader().setFixedHeight(_DLG_HDR_H)
        for r in range(tbl.rowCount()):
            tbl.setRowHeight(r, _DLG_ROW_H)
        self._tbl_layout.addWidget(tbl)

    def _export_both(self):
        import os, tempfile
        from data.processor import get_ecu_status, get_ecu_serienumber

        try:
            df1 = get_ecu_status(self._aircraft, '1', self._df_cyc, self._sys_vars)
            df2 = get_ecu_status(self._aircraft, '2', self._df_cyc, self._sys_vars)
            sn1 = get_ecu_serienumber(self._aircraft, '1', self._df_cfg) if self._df_cfg is not None else 'n/b'
            sn2 = get_ecu_serienumber(self._aircraft, '2', self._df_cfg) if self._df_cfg is not None else 'n/b'
        except Exception as exc:
            QMessageBox.warning(self, 'Export error', str(exc))
            return

        def _tab(ecu_nr, sn):
            return f'{self._aircraft} ECU{ecu_nr} SN {sn}'[:31]

        try:
            tmp = tempfile.NamedTemporaryFile(
                suffix='.xlsx', prefix=f'ECU_{self._aircraft}_', delete=False)
            tmp.close()
            with pd.ExcelWriter(tmp.name, engine='openpyxl') as writer:
                df1.to_excel(writer, index=False, sheet_name=_tab('1', sn1))
                df2.to_excel(writer, index=False, sheet_name=_tab('2', sn2))
                _style_excel_sheet(writer.sheets[_tab('1', sn1)], df1)
                _style_excel_sheet(writer.sheets[_tab('2', sn2)], df2)
            os.startfile(tmp.name)
        except Exception as exc:
            QMessageBox.warning(self, 'Export error', str(exc))


# ---------------------------------------------------------------------------
# Tabel-bouwer
# ---------------------------------------------------------------------------

def _build_table(df: pd.DataFrame, columns: list, headers: list,
                 col_widths: list, rest_col: str | None,
                 max_rows: int = N_ROWS, max_stretch: int = 0) -> QTableWidget:
    rows = min(len(df), max_rows)
    tbl  = QTableWidget(rows, len(columns))
    tbl.setHorizontalHeaderLabels(headers)
    tbl.setAlternatingRowColors(False)   # handmatig gezet zodat NEG_BG niet wordt overschreven
    tbl.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
    tbl.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
    tbl.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
    tbl.verticalHeader().setVisible(False)
    tbl.setShowGrid(True)
    tbl.setStyleSheet(_TBL_QSS)
    tbl.setItemDelegate(_BgDelegate(tbl))
    tbl.setFixedHeight(TBL_H)
    tbl.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Fixed)

    hdr = tbl.horizontalHeader()
    hdr.setHighlightSections(False)
    hdr.setFixedHeight(HDR_H)
    hdr.setDefaultAlignment(Qt.AlignmentFlag.AlignCenter)

    stretch_col = -1
    for i, (mode, width) in enumerate(col_widths):
        if mode == 'F':
            hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Fixed)
            tbl.setColumnWidth(i, width)
        elif mode == 'S':
            hdr.setSectionResizeMode(i, QHeaderView.ResizeMode.Stretch)
            stretch_col = i

    if max_stretch > 0 and stretch_col >= 0:
        class _StretchCapFilter(QObject):
            def eventFilter(self_, obj, event):
                if event.type() == QEvent.Type.Resize:
                    fixed = sum(
                        tbl.columnWidth(j) for j in range(tbl.columnCount())
                        if j != stretch_col
                    )
                    avail = tbl.viewport().width() - fixed
                    w = max(40, min(avail, max_stretch))
                    hdr.setSectionResizeMode(stretch_col, QHeaderView.ResizeMode.Fixed)
                    tbl.setColumnWidth(stretch_col, w)
                    hdr.setSectionResizeMode(stretch_col, QHeaderView.ResizeMode.Stretch)
                return False
        tbl._stretch_cap = _StretchCapFilter(tbl)
        tbl.installEventFilter(tbl._stretch_cap)

    row_bg = [QColor(TBL_BG), QColor(TBL_ALT)]
    for r, (_, row) in enumerate(df.head(max_rows).iterrows()):
        tbl.setRowHeight(r, ROW_H)
        for c, col in enumerate(columns):
            val  = row.get(col, '')
            text = '' if pd.isna(val) else str(val)
            item = QTableWidgetItem(text)
            item.setFlags(Qt.ItemFlag.ItemIsEnabled | Qt.ItemFlag.ItemIsSelectable)
            if col == rest_col:
                item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
                cell_neg = False
                try:
                    cell_neg = float(str(val)) < 0
                except (ValueError, TypeError):
                    pass
                if cell_neg:
                    item.setForeground(QBrush(QColor(NEG_FG)))
                    item.setBackground(QBrush(QColor(NEG_BG)))
                else:
                    item.setBackground(QBrush(row_bg[r % 2]))
            else:
                item.setBackground(QBrush(row_bg[r % 2]))
            tbl.setItem(r, c, item)
    return tbl


# ---------------------------------------------------------------------------
# Info-kaart links van de tabellen
# ---------------------------------------------------------------------------

def _build_info_card(aircraft: str, hrs: float,
                     df_cal: pd.DataFrame, df_cyc: pd.DataFrame, df_cfg,
                     user_vars: dict, sys_vars: dict) -> QFrame:
    from data.processor import get_serienummers, get_tellerstanden

    card = QFrame()
    card.setFixedWidth(INFO_W)
    card.setFixedHeight(TBL_H)
    card.setStyleSheet(f"""
        QFrame {{
            background-color: {AC_CARD_COLOR};
            border-right: 1px solid {AC_CARD_BDR};
        }}
    """)
    v = QVBoxLayout(card)
    v.setContentsMargins(6, 6, 6, 6)
    v.setSpacing(2)
    v.setAlignment(Qt.AlignmentFlag.AlignTop)

    # Kist-nummer
    lbl_name = QLabel(aircraft)
    f = QFont()
    f.setBold(True)
    f.setPointSize(14)
    lbl_name.setFont(f)
    lbl_name.setStyleSheet('color: white; background: transparent;')
    v.addWidget(lbl_name)

    # Vlieguren
    hrs_str = f'{hrs:,.1f}'.replace(',', '')
    lbl_hrs = QLabel(f'{hrs_str} hrs')
    lbl_hrs.setStyleSheet('color: #ffffff; font-size: 13px; background: transparent;')
    v.addWidget(lbl_hrs)

    v.addSpacing(6)

    # Knoppen
    def open_sn():
        if df_cfg is not None:
            df  = get_serienummers(aircraft, df_cfg, sys_vars)
            dlg = _InfoDialog(f'Serial numbers  -  {aircraft}', df,
                              [('F', 210), ('S', 0)], card.window())
            dlg.exec()

    def open_hrs():
        df  = get_tellerstanden(aircraft, df_cyc, sys_vars)
        dlg = _InfoDialog(f'Item hours  -  {aircraft}', df,
                          [('S', 0), ('F', 110), ('F', 90)], card.window())
        dlg.exec()

    def open_ecu():
        dlg = _EcuDialog(aircraft, df_cyc, df_cfg, sys_vars, card.window())
        dlg.exec()

    def open_pdf():
        import traceback
        try:
            from export.pdf_export import create_atl_pdf
            create_atl_pdf(aircraft, df_cal, df_cyc, user_vars, sys_vars)
        except Exception:
            QMessageBox.warning(card.window(), 'ATL Report error',
                                traceback.format_exc())

    def open_specials():
        # Laad vlak voor openen de nieuwste user_vars van schijf
        # zodat de dialoog niet afhankelijk is van de 5s polling-cyclus.
        try:
            from data.processor import load_user_variables
            latest_user_vars = load_user_variables()
        except Exception:
            latest_user_vars = user_vars
        dlg = _SpecialsDialog(aircraft, latest_user_vars, card.window())
        if dlg.exec() and hasattr(card.window(), '_refresh_overview'):
            card.window()._refresh_overview()

    for label, fn in [
        ('ATL Report', open_pdf),
        ('Specials',   open_specials),
    ]:
        btn = QPushButton(label)
        btn.setStyleSheet(_BTN_QSS)
        btn.setFixedWidth(INFO_W - 12)
        btn.clicked.connect(fn)
        v.addWidget(btn)

    v.addSpacing(4)

    for label, fn in [
        ('Serial Nrs', open_sn),
        ('Item Hours', open_hrs),
        ('ECU',        open_ecu),
    ]:
        btn = QPushButton(label)
        btn.setStyleSheet(_BTN_QSS)
        btn.setFixedWidth(INFO_W - 12)
        btn.clicked.connect(fn)
        v.addWidget(btn)

    return card


# ---------------------------------------------------------------------------
# Sectie per vliegtuig
# ---------------------------------------------------------------------------

def _build_ac_section(aircraft: str, df_cal: pd.DataFrame, df_cyc: pd.DataFrame,
                       df_cfg, user_vars: dict, sys_vars: dict,
                       fh_poref: str) -> QWidget:
    from data.processor import get_ac_hrs, get_bijzonderheden

    hrs = get_ac_hrs(df_cyc, aircraft, fh_poref)

    # Buitenste frame geeft het "blok" uiterlijk
    block = QFrame()
    block.setStyleSheet(f"""
        QFrame {{
            background: {TBL_BG};
            border: 1px solid {AC_HDR_BDR};
            border-radius: 3px;
        }}
    """)
    v = QVBoxLayout(block)
    v.setContentsMargins(0, 0, 0, 0)
    v.setSpacing(0)

    # Tabel-rij
    row = QWidget()
    row.setStyleSheet('background: transparent;')
    h = QHBoxLayout(row)
    h.setContentsMargins(0, 0, 0, 0)
    h.setSpacing(0)

    # Info-kaart (naam + uren + knoppen)
    h.addWidget(_build_info_card(aircraft, hrs, df_cal, df_cyc, df_cfg, user_vars, sys_vars))
    _add_div(h)

    # Kalender
    df_ac_cal = df_cal[df_cal['Aircraft'] == aircraft]
    h.addWidget(_build_table(
        df_ac_cal,
        ['POplan', 'PoRef', 'PoRef Omschrijving', 'Rest'],
        ['Po Plan', 'Po Ref', 'Omschrijving', 'Rest\ndag'],
        [('F', POPLAN_W), ('F', POREF_W), ('S', 0), ('F', 46)],
        rest_col='Rest', max_stretch=OMSCHR_MAX,
    ), stretch=3)
    _add_div(h)

    # Uren
    df_ac_hrs = df_cyc[
        (df_cyc['Aircraft'] == aircraft) &
        df_cyc['Kenmerknaam'].str.contains('HOURS', na=False)
    ]
    h.addWidget(_build_table(
        df_ac_hrs,
        ['POplan', 'PoRef', 'PoRef Omschrijving', 'Rest'],
        ['Po Plan', 'Po Ref', 'Omschrijving', 'Rest\nUren'],
        [('F', POPLAN_W), ('F', POREF_W), ('S', 0), ('F', 46)],
        rest_col='Rest', max_stretch=OMSCHR_MAX,
    ), stretch=3)
    _add_div(h)

    # Cyclus
    df_ac_cyc = df_cyc[
        (df_cyc['Aircraft'] == aircraft) &
        ~df_cyc['Kenmerknaam'].str.contains('HOURS', na=False)
    ].copy()
    df_ac_cyc['Kenmerknaam'] = df_ac_cyc['Kenmerknaam'].str.lower()
    h.addWidget(_build_table(
        df_ac_cyc,
        ['POplan', 'PoRef', 'PoRef Omschrijving', 'Rest', 'Kenmerknaam'],
        ['Po Plan', 'Po Ref', 'Omschrijving', 'Rest', 'KenmerkNaam'],
        [('F', POPLAN_W), ('F', POREF_W), ('S', 0), ('F', 46), ('F', 102)],
        rest_col='Rest', max_stretch=OMSCHR_MAX,
    ), stretch=4)
    _add_div(h)

    # Bijzonderheden
    df_bijz = get_bijzonderheden(aircraft, user_vars, sys_vars, df_cyc)
    h.addWidget(_build_table(
        df_bijz,
        ['Bijzonderheid', 'PoPlan', 'Rest', 'Eenheid'],
        ['Bijzonderheid', 'PO', 'Rest', 'Eenheid'],
        [('S', 0), ('F', 44), ('F', 46), ('F', 84)],
        rest_col='Rest',
    ), stretch=3)

    v.addWidget(row)

    # Wikkel-widget met marge voor het blok-effect
    wrapper = QWidget()
    wl = QVBoxLayout(wrapper)
    wl.setContentsMargins(2, 3, 6, 3)
    wl.setSpacing(0)
    wl.addWidget(block)
    return wrapper


def _add_div(layout: QHBoxLayout) -> None:
    sep = QFrame()
    sep.setFixedWidth(6)
    sep.setStyleSheet(f'background: {SLATE_900}; border: none;')
    layout.addWidget(sep)


# ---------------------------------------------------------------------------
# Hoofd-widget
# ---------------------------------------------------------------------------

class OverviewTab(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        layout = QVBoxLayout(self)
        layout.setContentsMargins(0, 0, 0, 0)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet(f'QScrollArea {{ background: {SLATE_900}; border: none; }}')
        scroll.viewport().setStyleSheet(f'background: {SLATE_900};')

        self._container = QWidget()
        self._container.setStyleSheet(f'background: {SLATE_900};')
        self._con_layout = QVBoxLayout(self._container)
        self._con_layout.setContentsMargins(0, 0, 0, 0)
        self._con_layout.setSpacing(0)
        self._con_layout.addStretch()

        scroll.setWidget(self._container)
        layout.addWidget(scroll)

    def load_data(self, store, sys_vars: dict, user_vars: dict) -> None:
        from data.processor import (
            get_aircraft_list, get_calendar_inspections,
            get_cycle_inspections, prepare_configuratie, prepare_statusbord,
        )

        if store.statusbord is None:
            return

        df_sb  = prepare_statusbord(store.statusbord)
        df_cal = get_calendar_inspections(df_sb)
        df_cyc = get_cycle_inspections(df_sb)
        df_cfg = prepare_configuratie(store.configuratie) if store.configuratie is not None else None

        aircraft_list = get_aircraft_list(df_sb, user_vars)
        fh_poref = sys_vars.get('Kenmerken', {}).get('Flight Hrs', [''])[0]

        while self._con_layout.count():
            item = self._con_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        for ac in aircraft_list:
            section = _build_ac_section(
                ac, df_cal, df_cyc, df_cfg, user_vars, sys_vars, fh_poref)
            self._con_layout.addWidget(section)

        self._con_layout.addStretch()


def get_bijzonderheden(aircraft, user_vars, sys_vars, df_cyc):
    from data.processor import get_bijzonderheden as _g
    return _g(aircraft, user_vars, sys_vars, df_cyc)
